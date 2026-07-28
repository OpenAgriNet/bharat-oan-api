"""Build the results workbook from judged scores.

Produces a workbook laid out like the Gemma comparison file: a Summary tab
with the seven headline metrics averaged per language, a per-question results tab,
and a per-category tab. Where the Gemma run covered the same session_id, its score
is shown alongside for reference.

Usage:
    .venv/bin/python3 -m benchmarks.voice_quality.report
"""

import argparse
import json
import statistics
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent / "data"
GEMMA_WORKBOOK = Path.home() / "Downloads" / "gemmacomparison.xlsx"

LANG_NAMES = {
    "as": "assamese", "bn": "bengali", "en": "english", "gu": "gujarati", "hi": "hindi",
    "kn": "kannada", "ml": "malayalam", "mr": "marathi", "ta": "tamil", "te": "telugu",
}

# The seven metrics the Summary tab of the reference workbook reports, with their maxima.
HEADLINE = [
    ("tool_call_quality", "Tool call quality", 4),
    ("translation_accuracy", "Translation quality", 10),
    ("accuracy_completeness", "Accuracy_completeness", 4),
    ("brevity", "Brevity", 4),
    ("no_fabrication", "No_Fabrication", 1),
    ("grammar_fluency", "Grammar_Fluency", 4),
    ("language_purity", "Language purity", 4),
]

SECONDARY = [
    ("actionability", "Actionability", 1),
    ("conversation_closure", "Conversation_Closure", 1),
    ("source_data_comprehensiveness", "Source_Data_Comprehensiveness", 1),
    ("citation_accuracy", "Citation_Accuracy", 1),
    ("citation_comprehensiveness", "Citation_Comprehensiveness", 4),
    ("output_hygiene", "Output_Hygiene", 1),
]

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
BOLD = Font(bold=True)


def mean(values) -> float | None:
    clean = [v for v in values if isinstance(v, (int, float))]
    return statistics.mean(clean) if clean else None


def load_gemma_baseline() -> dict[tuple[str, int], dict[str, float]]:
    """{(lang_code, session_id): {metric: score}} from the reference workbook."""
    if not GEMMA_WORKBOOK.exists():
        return {}
    name_to_code = {v: k for k, v in LANG_NAMES.items()}
    workbook = openpyxl.load_workbook(GEMMA_WORKBOOK, data_only=True, read_only=True)
    sheet = workbook["gemma results"]
    rows = sheet.iter_rows(values_only=True)
    header = list(next(rows))
    file_i, session_i = header.index("file"), header.index("session_id")
    score_cols = {h[len("score_"):]: i for i, h in enumerate(header)
                  if isinstance(h, str) and h.startswith("score_")}

    baseline: dict[tuple[str, int], dict[str, float]] = {}
    for row in rows:
        if not row or len(row) <= session_i or not row[file_i] or row[session_i] is None:
            continue
        code = name_to_code.get(str(row[file_i]).split("_")[0])
        if code is None:
            continue
        baseline[(code, int(row[session_i]))] = {
            m: row[i] for m, i in score_cols.items()
            if i < len(row) and isinstance(row[i], (int, float))
        }
    workbook.close()
    return baseline


def style_header(sheet, row: int = 1) -> None:
    for cell in sheet[row]:
        if cell.value is not None:
            cell.font = BOLD
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(sheet, max_width: int = 55) -> None:
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        longest = max((len(str(c.value)) for c in column if c.value is not None), default=8)
        sheet.column_dimensions[letter].width = min(max(10, longest + 2), max_width)


def write_summary(sheet, scores: list[dict], baseline: dict, meta: dict) -> None:
    sheet.append(["Bharat Vistaar voice quality - LLM-as-judge"])
    sheet["A1"].font = Font(bold=True, size=13)
    sheet.append([f"Model under test: {meta['model_under_test']}"])
    sheet.append([f"Judge: {meta['judge']}"])
    sheet.append([f"{meta['per_language']} questions per language, {meta['n_items']} total"
                  f" | random sample, seed {meta['seed']}"])
    sheet.append([])

    header_row = sheet.max_row + 1
    sheet.append(["Language"] + [f"{label} (max {mx})" for _, label, mx in HEADLINE]
                 + ["Gemma tool call", "Gemma translation", "n"])
    style_header(sheet, header_row)

    by_lang: dict[str, list[dict]] = defaultdict(list)
    for row in scores:
        by_lang[row["language"]].append(row)

    for code, name in LANG_NAMES.items():
        rows = by_lang.get(code, [])
        if not rows:
            continue
        sids = [r["session_id"] for r in rows]
        line = [name]
        for metric, _, _ in HEADLINE:
            line.append(mean(r.get(f"score_{metric}") for r in rows))
        line.append(mean(baseline.get((code, s), {}).get("tool_call_quality") for s in sids))
        line.append(mean(baseline.get((code, s), {}).get("translation_accuracy") for s in sids))
        line.append(len(rows))
        sheet.append(line)

    overall = ["ALL LANGUAGES"]
    for metric, _, _ in HEADLINE:
        overall.append(mean(r.get(f"score_{metric}") for r in scores))
    all_sids = [(r["language"], r["session_id"]) for r in scores]
    overall.append(mean(baseline.get(k, {}).get("tool_call_quality") for k in all_sids))
    overall.append(mean(baseline.get(k, {}).get("translation_accuracy") for k in all_sids))
    overall.append(len(scores))
    sheet.append(overall)
    for cell in sheet[sheet.max_row]:
        cell.font = BOLD

    sheet.append([])
    secondary_row = sheet.max_row + 1
    sheet.append(["Language"] + [f"{label} (max {mx})" for _, label, mx in SECONDARY]
                 + ["Avg seconds", "Avg words"])
    style_header(sheet, secondary_row)
    for code, name in LANG_NAMES.items():
        rows = by_lang.get(code, [])
        if not rows:
            continue
        line = [name] + [mean(r.get(f"score_{m}") for r in rows) for m, _, _ in SECONDARY]
        line.append(mean(r.get("score_elapsed_seconds") for r in rows))
        line.append(mean(r.get("score_word_count") for r in rows))
        sheet.append(line)

    for row in sheet.iter_rows(min_row=header_row + 1):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.00"


def write_results(sheet, scores: list[dict], baseline: dict) -> None:
    metrics = [m for m, _, _ in HEADLINE] + [m for m, _, _ in SECONDARY]
    header = (
        ["language", "session_id", "category", "question", "question_en", "answer",
         "tool_calls", "turn_count", "date_clarification_needed"]
        + [f"score_{m}" for m in metrics]
        + ["score_elapsed_seconds", "score_word_count", "score_error"]
        + [f"gemma_score_{m}" for m, _, _ in HEADLINE]
        + [f"reason_{m}" for m in metrics]
        + ["judge_error"]
    )
    sheet.append(header)
    style_header(sheet)
    sheet.freeze_panes = "A2"

    for row in sorted(scores, key=lambda r: (r["language"], r["session_id"])):
        gemma = baseline.get((row["language"], row["session_id"]), {})
        sheet.append(
            [row["language"], row["session_id"], row["category"], row["question"],
             row["question_en"], row["answer"], row["tool_calls"], row["turn_count"],
             row["date_clarification_needed"]]
            + [row.get(f"score_{m}") for m in metrics]
            + [row.get("score_elapsed_seconds"), row.get("score_word_count"), row.get("score_error")]
            + [gemma.get(m) for m, _, _ in HEADLINE]
            + [row.get(f"reason_{m}") for m in metrics]
            + [row.get("judge_error")]
        )


def write_by_category(sheet, scores: list[dict]) -> None:
    sheet.append(["Category", "n"] + [f"{label} (max {mx})" for _, label, mx in HEADLINE]
                 + ["Tool call rate", "Avg seconds"])
    style_header(sheet)

    by_cat: dict[str, list[dict]] = defaultdict(list)
    for row in scores:
        by_cat[row["category"]].append(row)

    for category, rows in sorted(by_cat.items(), key=lambda kv: -len(kv[1])):
        line = [category, len(rows)]
        line += [mean(r.get(f"score_{m}") for r in rows) for m, _, _ in HEADLINE]
        line.append(sum(1 for r in rows if r["tool_calls"]) / len(rows))
        line.append(mean(r.get("score_elapsed_seconds") for r in rows))
        sheet.append(line)

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.00"


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--scores", default=str(DATA_DIR / "scores.json"))
    parser.add_argument("--model", default="gpt-5.4-mini")
    parser.add_argument("--judge", default="gpt-4.1")
    parser.add_argument("--out", default=str(DATA_DIR.parent / "bh_voice_quality_gpt54mini.xlsx"))
    args = parser.parse_args()

    scores = [r for r in json.loads(Path(args.scores).read_text(encoding="utf-8")) if r]
    dataset = json.loads((DATA_DIR / "dataset.json").read_text(encoding="utf-8"))
    baseline = load_gemma_baseline()

    meta = {
        "model_under_test": args.model,
        "judge": args.judge,
        "per_language": dataset["per_language"],
        "n_items": len(scores),
        "seed": dataset["seed"],
    }

    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "Summary"
    write_summary(summary, scores, baseline, meta)
    write_by_category(workbook.create_sheet("By category"), scores)
    write_results(workbook.create_sheet("Results"), scores, baseline)

    for sheet in workbook.worksheets:
        autosize(sheet)

    out_path = Path(args.out)
    workbook.save(out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
