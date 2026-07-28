"""Build the two-tab benchmark marksheet for the model under test.

Tab 1 "Summary"  - the seven headline metrics, averaged per language, for this model
                   only (no Gemma comparison columns).
Tab 2 "<model> results" - one row per question, using the same column order as the
                   "gemma results" sheet in the Gemma comparison workbook so
                   the two can be diffed or stacked directly.

Usage:
    .venv/bin/python3 -m benchmarks.voice_quality.marksheet
"""

import argparse
import json
import statistics
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA_DIR = Path(__file__).resolve().parent / "data"

LANG_NAMES = {
    "as": "assamese", "bn": "bengali", "en": "english", "gu": "gujarati", "hi": "hindi",
    "kn": "kannada", "ml": "malayalam", "mr": "marathi", "ta": "tamil", "te": "telugu",
}

# The seven metrics the reference workbook's Summary tab reports, in its order.
HEADLINE = [
    ("tool_call_quality", "Tool call quality ( max: 4)"),
    ("translation_accuracy", "Translation quality (max: 10)"),
    ("accuracy_completeness", "Accuracy_completeness  ( max: 4)"),
    ("brevity", "Brevity  ( max: 4)"),
    ("no_fabrication", "No_Fabrication  ( max: 1)"),
    ("grammar_fluency", "Grammar_Fluency  ( max: 4)"),
    ("language_purity", "Language purity  ( max: 4)"),
]

# Column order copied from the "gemma results" sheet.
RESULTS_COLUMNS = [
    "file", "session_id", "questions_answers", "question", "answer", "tool_calls",
    "score_translation_accuracy", "score_tool_call_quality",
    "reason_translation_accuracy", "reason_tool_call_quality",
    "score_accuracy_completeness", "score_actionability", "score_conversation_closure",
    "score_source_data_comprehensiveness", "score_no_fabrication", "score_citation_accuracy",
    "score_citation_comprehensiveness", "score_grammar_fluency", "score_language_purity",
    "score_brevity", "score_output_hygiene", "score_elapsed_seconds", "score_word_count",
    "score_error",
    "reason_accuracy_completeness", "reason_actionability", "reason_conversation_closure",
    "reason_source_data_comprehensiveness", "reason_no_fabrication", "reason_citation_accuracy",
    "reason_citation_comprehensiveness", "reason_grammar_fluency", "reason_language_purity",
    "reason_brevity", "reason_output_hygiene", "reason_elapsed_seconds", "reason_word_count",
    "reason_error",
]

# Mechanical columns are computed, not judged; these mirror the reference sheet's wording.
MECHANICAL_REASONS = {
    "reason_output_hygiene": "No internal artifacts",
    "reason_elapsed_seconds": "Wall-clock seconds for the full pipeline run",
    "reason_word_count": "Computed from output text",
    "reason_error": "No error present",
}

HEADER_FILL = PatternFill("solid", fgColor="DDEBF7")
BOLD = Font(bold=True)


def mean(values) -> float | None:
    clean = [v for v in values if isinstance(v, (int, float))]
    return statistics.mean(clean) if clean else None


def style_header(sheet, row: int = 1) -> None:
    for cell in sheet[row]:
        if cell.value is not None:
            cell.font = BOLD
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(sheet, max_width: int = 50) -> None:
    for column in sheet.columns:
        letter = get_column_letter(column[0].column)
        longest = max((len(str(c.value)) for c in column[:60] if c.value is not None), default=8)
        sheet.column_dimensions[letter].width = min(max(10, longest + 2), max_width)


def build_qa(record: dict) -> str:
    """Render the conversation the way the reference sheet's questions_answers column does."""
    parts = []
    turn = 0
    for entry in record["turns"]:
        if entry["role"] == "user":
            turn += 1
            parts.append(f"Q{turn}: {entry['text']}")
        else:
            parts.append(f"A{turn}: {entry['text']}")
    return "  ".join(parts)


def write_summary(sheet, scores: list[dict], meta: dict) -> None:
    sheet.append(["Langauge"] + [label for _, label in HEADLINE])
    style_header(sheet)
    sheet.freeze_panes = "B2"

    by_lang: dict[str, list[dict]] = defaultdict(list)
    for row in scores:
        by_lang[row["language"]].append(row)

    for code, name in LANG_NAMES.items():
        rows = by_lang.get(code, [])
        if not rows:
            continue
        sheet.append([name] + [mean(r.get(f"score_{m}") for r in rows) for m, _ in HEADLINE])

    sheet.append(["ALL LANGUAGES"] + [mean(r.get(f"score_{m}") for r in scores) for m, _ in HEADLINE])
    for cell in sheet[sheet.max_row]:
        cell.font = BOLD

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.00"

    sheet.append([])
    sheet.append([f"Model under test: {meta['model']}"])
    sheet.append([f"Judge: {meta['judge']}"])
    sheet.append([f"{meta['per_language']} questions per language, {meta['n']} total"
                  f" (random sample, seed {meta['seed']})"])


def write_results(sheet, scores: list[dict], responses: dict, model_tag: str) -> None:
    sheet.append(RESULTS_COLUMNS)
    style_header(sheet)
    sheet.freeze_panes = "C2"

    for row in sorted(scores, key=lambda r: (LANG_NAMES[r["language"]], r["session_id"])):
        record = responses[(row["language"], row["session_id"])]
        values = {
            "file": f"{LANG_NAMES[row['language']]}_{model_tag}",
            "session_id": row["session_id"],
            "questions_answers": build_qa(record),
            "question": row["question"],
            "answer": row["answer"],
            "tool_calls": row["tool_calls"],
            **MECHANICAL_REASONS,
        }
        if record["error"]:
            values["reason_error"] = record["error"]

        sheet.append([values.get(col, row.get(col)) for col in RESULTS_COLUMNS])

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.00"
            cell.alignment = Alignment(vertical="top")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--model", default="gpt-5.4-mini")
    parser.add_argument("--model-tag", default="gpt54mini", help="suffix used in the file column")
    parser.add_argument("--judge", default="gpt-4.1")
    parser.add_argument("--out", default=str(DATA_DIR.parent / "bh_54mini_voice_quality_marksheet.xlsx"))
    args = parser.parse_args()

    scores = [r for r in json.loads((DATA_DIR / "scores.json").read_text(encoding="utf-8")) if r]
    raw_responses = json.loads((DATA_DIR / "responses.json").read_text(encoding="utf-8"))
    responses = {(r["language"], r["session_id"]): r for r in raw_responses if r}
    dataset = json.loads((DATA_DIR / "dataset.json").read_text(encoding="utf-8"))

    meta = {
        "model": args.model,
        "judge": args.judge,
        "per_language": dataset["per_language"],
        "n": len(scores),
        "seed": dataset["seed"],
    }

    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "Summary"
    write_summary(summary, scores, meta)
    write_results(workbook.create_sheet(f"{args.model} results"), scores, responses, args.model_tag)

    for sheet in workbook.worksheets:
        autosize(sheet)

    out_path = Path(args.out)
    workbook.save(out_path)
    print(f"wrote {out_path}")


if __name__ == "__main__":
    main()
